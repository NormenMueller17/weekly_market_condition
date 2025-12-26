from __future__ import annotations

from typing import Dict, Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule


def _build_header_map(ws: Worksheet, header_row: int = 1) -> Dict[str, int]:
    """Map header name -> column index (1-based). Ignores empty headers."""
    header_map: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if isinstance(val, str) and val.strip():
            header_map[val.strip()] = col
    return header_map


def autofit_columns_by_header(
    ws: Worksheet,
    header_row: int = 1,
    min_width: int = 10,
    padding: int = 2,
    max_width: int = 60,
) -> None:
    """Set column widths based on header text length (not data)."""
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=header_row, column=col).value
        header_txt = (str(header_val) if header_val is not None else "").strip()
        if not header_txt:
            continue
        width = min(max(len(header_txt) + padding, min_width), max_width)
        col_letter = ws.cell(row=header_row, column=col).column_letter
        ws.column_dimensions[col_letter].width = width


def apply_number_formats(
    ws: Worksheet,
    formats_by_colname: Dict[str, str],
    header_row: int = 1,
    data_start_row: int = 2,
) -> None:
    """Apply Excel number formats to columns by header name."""
    header_map = _build_header_map(ws, header_row=header_row)

    for col_name, fmt in formats_by_colname.items():
        col_idx = header_map.get(col_name)
        if not col_idx:
            continue
        for r in range(data_start_row, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = fmt


def sort_sheet_by_column(
    ws: Worksheet,
    sort_colname: str,
    header_row: int = 1,
    data_start_row: int = 2,
    ascending: bool = True,
) -> None:
    """Sort the sheet rows (in-place) by a column name."""
    header_map = _build_header_map(ws, header_row=header_row)
    sort_col_idx = header_map.get(sort_colname)
    if not sort_col_idx:
        return

    rows = []
    for r in range(data_start_row, ws.max_row + 1):
        row_values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        sort_key = ws.cell(row=r, column=sort_col_idx).value
        rows.append((sort_key, row_values))

    def _key(item):
        k = item[0]
        return (k is None, k)

    rows.sort(key=_key, reverse=not ascending)

    for i, (_, row_values) in enumerate(rows, start=data_start_row):
        for c, v in enumerate(row_values, start=1):
            ws.cell(row=i, column=c).value = v


def format_sheet(
    ws: Worksheet,
    formats_by_colname: Optional[Dict[str, str]] = None,
    autofit_headers: bool = True,
    sort_by: Optional[str] = None,
    sort_ascending: bool = True,
    header_row: int = 1,
    data_start_row: int = 2,
) -> None:
    """High-level helper: optional sort, optional number formats, optional autofit."""
    if sort_by:
        sort_sheet_by_column(
            ws,
            sort_colname=sort_by,
            header_row=header_row,
            data_start_row=data_start_row,
            ascending=sort_ascending,
        )

    if formats_by_colname:
        apply_number_formats(
            ws,
            formats_by_colname=formats_by_colname,
            header_row=header_row,
            data_start_row=data_start_row,
        )

    if autofit_headers:
        autofit_columns_by_header(ws, header_row=header_row)

def apply_formula_fills(
    ws: Worksheet,
    col_name: str,
    formula_fill_pairs: list[tuple[str, str]],
    header_row: int = 1,
    data_start_row: int = 2,
) -> None:
    """
    Apply conditional formatting (solid fills) to a column, addressed by header name.

    Parameters
    ----------
    col_name:
        Header text in row `header_row` to identify the column.
    formula_fill_pairs:
        List of (excel_formula, rgb_hex) pairs. rgb_hex is like "C6EFCE".
        Formulas should be written for the first data row (e.g. row 2) and
        include a fixed column reference like $AA2.
    """
    header_map = _build_header_map(ws, header_row=header_row)
    col_idx = header_map.get(col_name)
    if not col_idx or ws.max_row < data_start_row:
        return

    col_letter = ws.cell(row=header_row, column=col_idx).column_letter
    rng = f"{col_letter}{data_start_row}:{col_letter}{ws.max_row}"

    for formula, rgb in formula_fill_pairs:
        fill = PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")
        rule = FormulaRule(formula=[formula], fill=fill, stopIfTrue=True)
        ws.conditional_formatting.add(rng, rule)


def apply_debt_eps_conditional_formatting(
    ws: Worksheet,
    debt_col: str,
    eps_col: str,
    debt_thr_low: float = 0.5,
    debt_thr_med: float = 1.0,
    debt_thr_high: float = 2.0,
    eps_thr_strong: float = 10.0,
    eps_thr_mild: float = 3.0,
    eps_thr_flat: float = 3.0,
    header_row: int = 1,
    data_start_row: int = 2,
) -> None:
    """
    Apply "Ampel" background colors for:
      - Debt to Equity (lower is better)
      - EPS Acceleration (pp) (higher is better; neutral band around 0)

    Colors are simple defaults; feel free to change RGBs later.
    """
    # Default fills (Excel-like)
    GREEN = "C6EFCE"
    LIGHT_GREEN = "E2F0D9"
    YELLOW = "FFEB9C"
    ORANGE = "F8CBAD"
    RED = "FFC7CE"
    GREY = "D9D9D9"

    # Debt to Equity: <=low green, (low..med] yellow, (med..high] orange, >high red, non-number grey
    # We'll reference row 2 in formulas; openpyxl will apply across range.
    def _col_letter(name: str) -> str:
        hm = _build_header_map(ws, header_row=header_row)
        idx = hm.get(name)
        if not idx:
            return ""
        return ws.cell(row=header_row, column=idx).column_letter

    debt_letter = _col_letter(debt_col)
    eps_letter = _col_letter(eps_col)

    if debt_letter:
        pairs = [
            (f"=NOT(ISNUMBER(${debt_letter}2))", GREY),
            (f"=AND(ISNUMBER(${debt_letter}2),${debt_letter}2<={debt_thr_low})", GREEN),
            (f"=AND(ISNUMBER(${debt_letter}2),${debt_letter}2>{debt_thr_low},${debt_letter}2<={debt_thr_med})", YELLOW),
            (f"=AND(ISNUMBER(${debt_letter}2),${debt_letter}2>{debt_thr_med},${debt_letter}2<={debt_thr_high})", ORANGE),
            (f"=AND(ISNUMBER(${debt_letter}2),${debt_letter}2>{debt_thr_high})", RED),
        ]
        # Apply using apply_formula_fills but formula already references correct column letter
        apply_formula_fills(
            ws,
            debt_col,
            [(f, c) for f, c in pairs],
            header_row=header_row,
            data_start_row=data_start_row,
        )

    if eps_letter:
        # EPS Acceleration (pp):
        # >= strong green, [mild..strong) light green, (-flat..flat) yellow,
        # (-strong..-flat] orange, <= -strong red, non-number grey
        pairs = [
            (f"=NOT(ISNUMBER(${eps_letter}2))", GREY),
            (f"=AND(ISNUMBER(${eps_letter}2),${eps_letter}2>={eps_thr_strong})", GREEN),
            (f"=AND(ISNUMBER(${eps_letter}2),${eps_letter}2>={eps_thr_mild},${eps_letter}2<{eps_thr_strong})", LIGHT_GREEN),
            (f"=AND(ISNUMBER(${eps_letter}2),${eps_letter}2>-{eps_thr_flat},${eps_letter}2<{eps_thr_flat})", YELLOW),
            (f"=AND(ISNUMBER(${eps_letter}2),${eps_letter}2<=-{eps_thr_flat},${eps_letter}2>-{eps_thr_strong})", ORANGE),
            (f"=AND(ISNUMBER(${eps_letter}2),${eps_letter}2<=-{eps_thr_strong})", RED),
        ]
        apply_formula_fills(
            ws,
            eps_col,
            [(f, c) for f, c in pairs],
            header_row=header_row,
            data_start_row=data_start_row,
        )



def apply_industry_percentile_conditional_formatting(
    ws: Worksheet,
    metric_col: str,
    pctl_col: str,
    thr_top: float = 0.75,
    thr_mid: float = 0.50,
    thr_low: float = 0.25,
    header_row: int = 1,
    data_start_row: int = 2,
    hide_pctl_col: bool = True,
) -> None:
    """Color the *metric_col* cells based on a percentile column (0..1) in the same row.

    Typical thresholds:
      >=thr_top  -> GREEN
      >=thr_mid  -> LIGHT_GREEN
      >=thr_low  -> YELLOW
      < thr_low  -> RED
      non-numeric -> GREY
    """
    header_map = _build_header_map(ws, header_row=header_row)
    metric_idx = header_map.get(metric_col)
    pctl_idx = header_map.get(pctl_col)
    if not metric_idx or not pctl_idx:
        return

    metric_letter = ws.cell(row=header_row, column=metric_idx).column_letter
    pctl_letter = ws.cell(row=header_row, column=pctl_idx).column_letter

    # same palette as debt/eps
    GREY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    LIGHT_GREEN = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    pairs = [
        (f"=NOT(ISNUMBER(${pctl_letter}2))", GREY),
        (f"=AND(ISNUMBER(${pctl_letter}2),${pctl_letter}2>={thr_top})", GREEN),
        (f"=AND(ISNUMBER(${pctl_letter}2),${pctl_letter}2>={thr_mid},${pctl_letter}2<{thr_top})", LIGHT_GREEN),
        (f"=AND(ISNUMBER(${pctl_letter}2),${pctl_letter}2>={thr_low},${pctl_letter}2<{thr_mid})", YELLOW),
        (f"=AND(ISNUMBER(${pctl_letter}2),${pctl_letter}2<{thr_low})", RED),
    ]
    # apply fills to metric_col range, but formulas reference pctl column
    apply_formula_fills(
        ws,
        metric_col,
        [(f, c) for f, c in pairs],
        header_row=header_row,
        data_start_row=data_start_row,
    )

    if hide_pctl_col:
        ws.column_dimensions[pctl_letter].hidden = True
