import os
import sys
import json
import time
import pandas as pd
from datetime import datetime
from dataclasses import asdict
from config import SETTINGS
from data_sources import get_universe, get_company_info_map_from_csv, load_weekly_history, load_index_series
from breadth import compute_breadth, compute_breadth_snapshots_with_advancers as compute_breadth_snapshots, compute_sp500_breadth_200d
from emailer import send_email
from screener import screen_universe_minervini
from fetch_quote_data import batch_fetch_quote_data, fetch_quote_data_single
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import yfinance as yf
import warnings
from http_cache import try_enable_yfinance_cache, CacheConfig

# Silence noisy third-party warnings (optional)
warnings.filterwarnings("ignore", category=DeprecationWarning, module=r"yfinance\.scrapers\.fundamentals")
warnings.filterwarnings("ignore", category=FutureWarning, module=r"breadth")


from industry_strength import compute_industry_scores

from report_builder import (
    build_html_report,
    build_index_rows,
    build_risk_rows,
    build_sector_rows,
    heuristic_verdict,
)
from signal_generator import generate_signals, is_market_bullish, save_signals_json
import alpaca_client
import exit_manager
import trade_journal

# ── Test-Modus: TEST_MODE=1 → immer Test, TEST_MODE=0 → immer Live,
#    nicht gesetzt → nur samstags Live (weekday 5)
_tm_env = os.getenv("TEST_MODE", "")
if _tm_env == "1":
    TEST_MODE = True
elif _tm_env == "0":
    TEST_MODE = False
else:
    TEST_MODE = datetime.today().weekday() != 5

if TEST_MODE:
    print("[TEST-MODUS] Aktiv — keine Alpaca-Orders werden platziert oder storniert")

BOOLEAN_HEADERS = [
    "SMA10W steigend",
    "SMA30W steigend",
    "SMA40W steigend",
    "MA-Ordnung 10>30>40",
    "52W Range OK",
    "RS-Trend ↑",
    "Vol-Breakout",
    "Close > Vorwoche",
        ]

# -------------------------------------------------------------------
# Excel fill colors (central, easy to adjust)
# (These match typical Excel Conditional Formatting "Good/Neutral/Bad")
# -------------------------------------------------------------------
CF_GREEN_RGB      = "C6EFCE"  # Good (green)
CF_LIGHTGREEN_RGB = "EBF1DE"  # Mild positive
CF_YELLOW_RGB     = "FFEB9C"  # Neutral
CF_ORANGE_RGB     = "F8CBAD"  # Warning
CF_RED_RGB        = "FFC7CE"  # Bad (red)
CF_GRAY_RGB       = "E7E6E6"  # N/A

# Boolean Minervini columns
BOOL_TRUE_FILL_RGB  = CF_GREEN_RGB
BOOL_FALSE_FILL_RGB = CF_RED_RGB
BOOL_FONT_RGB       = "666666"

# Thresholds for the 5 highlighted metrics
DEBT_EQ_THR_LOW  = 0.50
DEBT_EQ_THR_MED  = 1.00
DEBT_EQ_THR_HIGH = 2.00

EPS_ACCEL_THR_STRONG = 10.0
EPS_ACCEL_THR_MILD   = 3.0
EPS_ACCEL_THR_FLAT   = 3.0  # +/- band

IND_PCTL_THR_TOP = 0.75
IND_PCTL_THR_MID = 0.50
IND_PCTL_THR_LOW = 0.25


from openpyxl.formatting.rule import FormulaRule

def _col_letter_by_header(ws, header_name: str, header_row: int = 1):
    header_map = {cell.value: cell.column for cell in ws[header_row] if cell.value}
    col_idx = header_map.get(header_name)
    if not col_idx:
        return None
    return get_column_letter(col_idx)

def _apply_cf_formula_fill(ws, cell_range: str, formula: str, rgb: str, stop: bool = True):
    fill = PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")
    rule = FormulaRule(formula=[formula], fill=fill, stopIfTrue=stop)
    ws.conditional_formatting.add(cell_range, rule)

def apply_debt_eps_conditional_formatting(ws, debt_header: str = "Debt to Equity", eps_header: str = "EPS Acceleration (pp)", start_row: int = 2):
    """Conditional formatting for Debt-to-Equity and EPS Acceleration."""
    debt_col = _col_letter_by_header(ws, debt_header)
    eps_col = _col_letter_by_header(ws, eps_header)

    if debt_col:
        rng = f"{debt_col}{start_row}:{debt_col}{ws.max_row}"
        c = debt_col
        r = start_row
        _apply_cf_formula_fill(ws, rng, f"NOT(ISNUMBER(${c}{r}))", CF_GRAY_RGB, stop=False)
        _apply_cf_formula_fill(ws, rng, f"AND(ISNUMBER(${c}{r}),${c}{r}<={DEBT_EQ_THR_LOW})", CF_GREEN_RGB)
        _apply_cf_formula_fill(ws, rng, f"AND(ISNUMBER(${c}{r}),${c}{r}>{DEBT_EQ_THR_LOW},${c}{r}<={DEBT_EQ_THR_MED})", CF_YELLOW_RGB)
        _apply_cf_formula_fill(ws, rng, f"AND(ISNUMBER(${c}{r}),${c}{r}>{DEBT_EQ_THR_MED},${c}{r}<={DEBT_EQ_THR_HIGH})", CF_ORANGE_RGB)
        _apply_cf_formula_fill(ws, rng, f"AND(ISNUMBER(${c}{r}),${c}{r}>{DEBT_EQ_THR_HIGH})", CF_RED_RGB)

    if eps_col:
        rng = f"{eps_col}{start_row}:{eps_col}{ws.max_row}"
        c = eps_col
        r = start_row
        _apply_cf_formula_fill(ws, rng, f"NOT(ISNUMBER(${c}{r}))", CF_GRAY_RGB, stop=False)
        _apply_cf_formula_fill(ws, rng, f"AND(ISNUMBER(${c}{r}),${c}{r}>={EPS_ACCEL_THR_STRONG})", CF_GREEN_RGB)
        _apply_cf_formula_fill(ws, rng, f"AND(ISNUMBER(${c}{r}),${c}{r}>={EPS_ACCEL_THR_MILD},${c}{r}<{EPS_ACCEL_THR_STRONG})", CF_LIGHTGREEN_RGB)
        _apply_cf_formula_fill(ws, rng, f"AND(ISNUMBER(${c}{r}),${c}{r}>-{EPS_ACCEL_THR_FLAT},${c}{r}<{EPS_ACCEL_THR_FLAT})", CF_YELLOW_RGB)
        _apply_cf_formula_fill(ws, rng, f"AND(ISNUMBER(${c}{r}),${c}{r}<=-{EPS_ACCEL_THR_FLAT},${c}{r}>-{EPS_ACCEL_THR_STRONG})", CF_ORANGE_RGB)
        _apply_cf_formula_fill(ws, rng, f"AND(ISNUMBER(${c}{r}),${c}{r}<=-{EPS_ACCEL_THR_STRONG})", CF_RED_RGB)

def apply_industry_percentile_conditional_formatting(ws, metric_header: str, pctl_header: str, start_row: int = 2, hide_pctl: bool = True):
    """Color metric cells based on helper percentile column (0..1)."""
    m_col = _col_letter_by_header(ws, metric_header)
    p_col = _col_letter_by_header(ws, pctl_header)
    if not m_col or not p_col:
        return

    rng = f"{m_col}{start_row}:{m_col}{ws.max_row}"
    c = p_col
    r = start_row
    _apply_cf_formula_fill(ws, rng, f"NOT(ISNUMBER(${c}{r}))", CF_GRAY_RGB, stop=False)
    _apply_cf_formula_fill(ws, rng, f"AND(ISNUMBER(${c}{r}),${c}{r}>={IND_PCTL_THR_TOP})", CF_GREEN_RGB)
    _apply_cf_formula_fill(ws, rng, f"AND(ISNUMBER(${c}{r}),${c}{r}>={IND_PCTL_THR_MID},${c}{r}<{IND_PCTL_THR_TOP})", CF_LIGHTGREEN_RGB)
    _apply_cf_formula_fill(ws, rng, f"AND(ISNUMBER(${c}{r}),${c}{r}>={IND_PCTL_THR_LOW},${c}{r}<{IND_PCTL_THR_MID})", CF_YELLOW_RGB)
    _apply_cf_formula_fill(ws, rng, f"AND(ISNUMBER(${c}{r}),${c}{r}<{IND_PCTL_THR_LOW})", CF_RED_RGB)

    if hide_pctl:
        # hide helper column
        header_map = {cell.value: cell.column for cell in ws[1] if cell.value}
        col_idx = header_map.get(pctl_header)
        if col_idx:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True



def style_boolean_columns(ws, headers=BOOLEAN_HEADERS, header_row: int = 1) -> None:
    """Färbt Bool-Spalten: WAHR/True -> grün, FALSCH/False -> rot; Text grau & zentriert."""
    header_to_col = {cell.value: cell.column for cell in ws[header_row] if cell.value}

    fill_green = PatternFill(fill_type="solid", fgColor=BOOL_TRUE_FILL_RGB)
    fill_red   = PatternFill(fill_type="solid", fgColor=BOOL_FALSE_FILL_RGB)
    font_gray  = Font(color=BOOL_FONT_RGB)
    center     = Alignment(horizontal="center", vertical="center")

    for head in headers:
        col = header_to_col.get(head)
        if not col:
            continue
        for row in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            sval = ("" if val is None else str(val)).strip().lower()
            is_true  = sval in ("true", "wahr", "1")
            is_false = sval in ("false", "falsch", "0")
            cell.font = font_gray
            cell.alignment = center
            if is_true:
                cell.fill = fill_green
            elif is_false:
                cell.fill = fill_red


    fill_green = PatternFill(fill_type="solid", fgColor="E6F4EA")  # hellgrün
    fill_red   = PatternFill(fill_type="solid", fgColor="FDE8E8")  # hellrot
    font_gray  = Font(color="666666")
    center     = Alignment(horizontal="center", vertical="center")

    for head in headers:
        col = header_to_col.get(head)
        if not col:
            continue
        for row in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            # robust: bool, WAHR/FALSCH, TRUE/FALSE, 1/0 …
            val = cell.value
            sval = ("" if val is None else str(val)).strip().lower()
            is_true  = sval in ("true", "wahr", "1")
            is_false = sval in ("false", "falsch", "0")
            cell.font = font_gray
            cell.alignment = center
            if is_true:
                cell.fill = fill_green
            elif is_false:
                cell.fill = fill_red
            else:
                # neutral: z.B. leere Zellen
                pass

PENDING_ORDERS_PATH = Path("docs") / "data" / "pending_orders.json"


def _save_pending_orders(
    top_picks:      list,
    sell_symbols:   list[str],
    sell_proceeds:  float,
    generated_date: str,
) -> None:
    """Serialize top-pick signals to pending_orders.json and wait for sell confirmation."""
    PENDING_ORDERS_PATH.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "generated_at": generated_date,
        "waiting_for_sells": [
            {
                "symbol":            sym,
                "approx_proceeds":   round(sell_proceeds / max(1, len(sell_symbols)), 2),
            }
            for sym in sell_symbols
        ],
        "signals": [asdict(s) for s in top_picks],
    }
    PENDING_ORDERS_PATH.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8"
    )


def monday_execute() -> None:
    """Place pending buy orders after confirming that the prerequisite sell orders filled.

    Run on Monday morning: python main.py --monday-execute
    """
    if not PENDING_ORDERS_PATH.exists():
        print("[MONDAY] Keine pending_orders.json gefunden – nichts zu tun.")
        return

    payload       = json.loads(PENDING_ORDERS_PATH.read_text(encoding="utf-8"))
    generated_at  = payload.get("generated_at", "")
    waiting_sells = payload.get("waiting_for_sells", [])
    signal_dicts  = payload.get("signals", [])
    sell_symbols  = [w["symbol"] for w in waiting_sells]

    print(f"[MONDAY] Pending Orders vom {generated_at}")
    print(f"[MONDAY] Warte auf Sells: {', '.join(sell_symbols)}")

    # 1. Sells bestätigen
    filled = alpaca_client.get_filled_sells_since(sell_symbols, generated_at)
    missing = [s for s in sell_symbols if s not in filled]
    if missing:
        print(f"[MONDAY] ⚠  Noch nicht gefüllt: {', '.join(missing)}")
        print("[MONDAY]    Bitte später erneut aufrufen oder manuell prüfen.")
        return

    for sym, fill in filled.items():
        print(
            f"[MONDAY] ✅ {sym} verkauft — "
            f"Qty {fill['qty']}  @${fill['filled_avg_price']:,.2f}  "
            f"({fill['filled_at']})"
        )

    # 2. Aktuelles Kapital holen
    portfolio = alpaca_client.get_portfolio()
    if portfolio is None:
        print("[MONDAY] ❌ Kein Alpaca-Portfolio – Abbruch.")
        return
    actual_cash = portfolio["cash"]
    print(f"[MONDAY] Aktuelles Cash: ${actual_cash:,.0f}")

    if actual_cash <= 0:
        print("[MONDAY] ⚠  Kein Cash verfügbar – keine Orders platziert.")
        return

    # 3. Buy-Orders mit realem Cash platzieren
    from signal_generator import TradeSignal
    placed_any = False
    for sig_dict in signal_dicts:
        try:
            sig = TradeSignal(**sig_dict)
        except Exception as e:
            print(f"[MONDAY] Signal-Deserialisierung fehlgeschlagen: {e}")
            continue

        # position_value auf tatsächlich verfügbares Cash begrenzen
        sig.position_value = min(sig.position_value, actual_cash)

        qty = int(sig.position_value // sig.buy_stop)
        if qty < 1:
            print(f"[MONDAY] ⏭  {sig.ticker}: qty=0 — Cash reicht nicht (${actual_cash:,.0f})")
            continue

        results = alpaca_client.place_signal_orders([sig], dry_run=TEST_MODE)
        for r in results:
            status = r["status"]
            if status == "placed":
                print(
                    f"[MONDAY] ✅ {sig.ticker}  Buy-Stop ${r['buy_stop']}  "
                    f"Max-Gap ${r['max_gap']}  Stop ${r['stop_loss']}  "
                    f"Qty {r['qty']}  → Order-ID {r['order_id']}"
                )
                placed_any = True
            elif status == "dry_run":
                print(f"[MONDAY] 🔍 DRY-RUN {sig.ticker}  qty={r['qty']}")
                placed_any = True
            else:
                print(f"[MONDAY] ❌ {sig.ticker}: {status}")

    if placed_any:
        PENDING_ORDERS_PATH.unlink(missing_ok=True)
        print("[MONDAY] pending_orders.json gelöscht.")


def run():
    #cache_enabled = try_enable_yfinance_cache(
    #CacheConfig(
    #    cache_name=".http_cache",
    #    expire_after_seconds=24 * 60 * 60,  # 24h TTL
    #    stale_if_error=True,
    #    )
    #)
    #if cache_enabled:
    #    print("[INFO] HTTP cache enabled for yfinance (.http_cache.sqlite)")

    report_date = datetime.now().strftime("%Y-%m-%d")
    # Cache aktivieren
    cache_config = CacheConfig(cache_name=".yfinance_cache", expire_after_seconds=24*60*60)
    cache_enabled = try_enable_yfinance_cache(cache_config)
    if cache_enabled:
        print("[CACHE] ✅ HTTP caching enabled (24h TTL)")

    
    # 1) Daten laden
    universe = get_universe()
    weekly = load_weekly_history(universe, weeks=SETTINGS.lookback_weeks)
    idx_data = load_index_series()

    print(f"[DEBUG] Universe size: {len(universe)}")
    non_empty = sum(1 for _, df in weekly.items() if isinstance(df, pd.DataFrame) and "Close" in df.columns and not df["Close"].dropna().empty)
    print(f"[DEBUG] Weekly non-empty datasets: {non_empty}")    

    # 2) Kennzahlen berechnen
    breadth_df   = compute_breadth(weekly)
    idx_rows     = build_index_rows(idx_data)
    risk_rows    = build_risk_rows(idx_data)
    sector_rows  = build_sector_rows(idx_data)

    # Market filter 1: S&P 500 10W EMA > 20W EMA
    market_bullish = is_market_bullish(idx_data.get("SPY"))
    print(f"[SIGNALS] Marktfilter 10EMA>20EMA: {'✅ BULLISH' if market_bullish else '❌ BÄRISCH – keine Kaufsignale'}")

    # Market filter 2: S&P 500 Marktbreite (% Aktien über 200d-MA)
    from signal_generator import _RULES_JSON as _rules_json
    _min_breadth = _rules_json.get("filters", {}).get("min_breadth_pct_200d", 40)
    sp500_breadth_pct = compute_sp500_breadth_200d()
    breadth_bullish = (_min_breadth == 0) or (sp500_breadth_pct >= _min_breadth)
    if not breadth_bullish:
        market_bullish = False
    print(f"[SIGNALS] S&P 500 Marktbreite: {sp500_breadth_pct:.1f}% über 200d {'✅' if breadth_bullish else f'❌ Kaufstopp (< {_min_breadth}%)'}")

    # 3) Report erzeugen
    summary = heuristic_verdict(breadth_df, idx_rows)
    report_date = pd.Timestamp.now().strftime("%Y-%m-%d")

    idx_df = pd.DataFrame.from_dict(dict(idx_rows), orient="index").T
    risk_df = pd.DataFrame(risk_rows, columns=["Metrik", "Aktuell", "Vorwoche", "Δ"]).set_index("Metrik")
    risk_df.rename(index={"TNX": "10Y Interest Rate (TNX)"}, inplace=True)
    risk_df.rename(index={"VIX": "Volatility Index (VIX)"}, inplace=True)
    risk_df.rename(index={"UUP": "US Dollar Index (UUP)"}, inplace=True)
    
    # Snapshots inkl. Advancers
    #breadth_snap = compute_breadth_snapshots(weekly)

    # Hintergrundfarben für alle Δ-Werte in risk_df
    def classify_delta(value: float, invert: bool = False) -> str:
        if pd.isna(value):
            return "neutral"
        if value > 0:
            return "neg" if invert else "pos"
        if value < 0:
            return "pos" if invert else "neg"
        return "neutral"
    
    risk_df["Δ_farbe"] = [
        classify_delta(delta, invert=("Volatility Index" in name))
        for name, delta in zip(risk_df.index, risk_df["Δ"])
    ]
    
    # 4) Marktführer nach Minervini screenen
    leaders = screen_universe_minervini(universe, min_score=0)
    info_map = get_company_info_map_from_csv()
    # NEU: Launchpad Quality Filter
    # Strenger Filter: Score ≥90 UND Range <8%
    if "Launchpad" in leaders.columns and "Launchpad Score" in leaders.columns:
        leaders.loc[
            (leaders["Launchpad"] == True) & 
            ((leaders["Launchpad Score"] < 90) | (leaders.get("Launchpad Range (%)", 100) >= 8)),
            "Launchpad"
        ] = False
        
        launchpad_count = len(leaders[leaders["Launchpad"] == True])
        print(f"[INFO] High-Quality Launchpads (Score >=80): {launchpad_count}")

    industry_table = pd.DataFrame()
    
    if not leaders.empty:
        # --- Sicherheitskopie (sehr wichtig für HTML-Formatierung später) ---
        leaders = leaders.copy()
    
        # Hilfsfunktion zum Entfernen von .DE
        def _make_sa_url(ticker: str) -> str:
            """
            Für US-Aktien:
                https://stockanalysis.com/stocks/TICKER
            Für deutsche (.DE):
                https://stockanalysis.com/quote/etr/TICKER (ohne .DE)
            """
            base = ticker.split(".")[0]
            if ticker.upper().endswith(".DE"):
                return f"https://stockanalysis.com/quote/etr/{base}"
            else:
                return f"https://stockanalysis.com/stocks/{base}"
                
        # Company & Industry (bereits geladen über info_map)
        leaders.insert(1, "Company", leaders.index.map(lambda t: info_map.get(t, {}).get("Company", "n/a")))
        leaders.insert(2, "SA", leaders.index.map(_make_sa_url))      
        leaders.insert(3, "Industry", leaders.index.map(lambda t: info_map.get(t, {}).get("Industry", "n/a")))

        # --- Fundamentaldaten nur für Leaders mit Score >= 6 laden ---
        # Score < 6 erscheinen nicht in Kaufsignalen und selten im Mail-Report (nur Score 8/8).
        # Damit werden 300-400 API-Calls auf ~20-50 reduziert.
        MIN_SCORE_FOR_FUNDAMENTALS = 6
        tickers_for_fundamentals = (
            leaders[pd.to_numeric(leaders["score"], errors="coerce") >= MIN_SCORE_FOR_FUNDAMENTALS]
            .index.tolist()
        )
        print(f"[INFO] Fundamentaldaten werden für {len(tickers_for_fundamentals)} "
              f"von {len(leaders)} Leaders geholt (Score >= {MIN_SCORE_FOR_FUNDAMENTALS}).")
        quote_map = batch_fetch_quote_data(tickers_for_fundamentals)
        leaders.insert(4, "Sektor", leaders.index.map(lambda t: quote_map.get(t, {}).get("Sector", "n/a")))
        leaders.insert(5, "Close", leaders.index.map(lambda t: quote_map.get(t, {}).get("Close")))
        # Fallback: Screener-Wochenschlusskurs wenn API-Fetch fehlschlug
        if "close_weekly_now" in leaders.columns:
            leaders["Close"] = leaders["Close"].combine_first(leaders["close_weekly_now"])
        leaders.insert(6, "MarketCap (Mio USD)", leaders.index.map(lambda t: quote_map.get(t, {}).get("MarketCap_Mio")))
        leaders.insert(7, "EPS (Forward/TTM)", leaders.index.map(lambda t: quote_map.get(t, {}).get("EPS_FWD_TTM")))
        leaders.insert(8, "EPS Wachstum FWD/TTM (%)", leaders.index.map(lambda t: quote_map.get(t, {}).get("EPS_GROWTH_FWD_TTM")))
        leaders.insert(9, "Revenue Wachstum TTM YoY (%)", leaders.index.map(lambda t: quote_map.get(t, {}).get("REV_GROWTH_TTM_YOY")))
        leaders.insert(10, "EPS Wachstum letztes Q YoY (%)", leaders.index.map(lambda t: quote_map.get(t, {}).get("EPS_GROWTH_LAST_Q_YOY")))

        "ROE (%)",
        "Operating Margin (%)",
        "FCF Margin (%)",
        "Debt to Equity",
        "EPS Acceleration (pp)",
        leaders.insert(10, "ROE (%)", leaders.index.map(lambda t: quote_map.get(t, {}).get("ROE")))
        leaders.insert(11, "Operating Margin (%)", leaders.index.map(lambda t: quote_map.get(t, {}).get("Operating_Margin")))
        leaders.insert(12, "FCF Margin (%)", leaders.index.map(lambda t: quote_map.get(t, {}).get("FCF_Margin")))
        leaders.insert(13, "Debt to Equity", leaders.index.map(lambda t: quote_map.get(t, {}).get("Debt_to_Equity")))
        leaders.insert(14, "EPS Acceleration (pp)", leaders.index.map(lambda t: quote_map.get(t, {}).get("EPS_Acceleration")))

      # Falls Screener noch keine 52W-Spalten liefert, zur Sicherheit anlegen
        if "52W High" not in leaders.columns:
            leaders["52W High"] = pd.NA
        if "Dist to 52W High (%)" not in leaders.columns:
            leaders["Dist to 52W High (%)"] = pd.NA

        # NEU: "Close Vorwoche" und "Veränderung in %"
        if "close_weekly_prev" in leaders.columns:
            leaders.insert(15, "Close Vorwoche", leaders["close_weekly_prev"])
        else:
            leaders.insert(15, "Close Vorwoche", pd.NA)

        if "close_weekly_change_pct" in leaders.columns:
            leaders.insert(16, "Veränderung in %", leaders["close_weekly_change_pct"])
        else:
            leaders.insert(16, "Veränderung in %", pd.NA)
        
        leaders.insert(17, "Ø-Volume 20T", leaders["vol20"])
        leaders.insert(18, "Volume Score", leaders["vol_score"])
        
        if "RS_delta_4w" in leaders.columns and "ΔRS 4W" not in leaders.columns:
            leaders["ΔRS 4W"] = leaders["RS_delta_4w"]
             
        # Alte Roh-Spalten nicht mehr gebraucht
        drop_cols = [
            "vol20",
            "vol_score",
            "close_weekly_now",
            "close_weekly_prev",
            "close_weekly_change_pct",
            "RS_now",
            "RS_4w",
            "RS_delta_4w",
        ]
        drop_cols = [c for c in drop_cols if c in leaders.columns]
        if drop_cols:
            leaders.drop(columns=drop_cols, inplace=True)

        # ------------------------------------------------------------
        # NEU: Industry Strength Scoring
        # Adds: Industry RS Score, Industry Strong Stock Score,
        #       Industry Volume Score (Activity*Direction, Variant 2),
        #       and the composite Industry Score.
        # ------------------------------------------------------------
        try:
            leaders, industry_table = compute_industry_scores(leaders)
            if industry_table is not None and not industry_table.empty:
                print(f"[INFO] Computed industry scores for {len(industry_table)} industries")
        except Exception as e:
            print(f"[WARN] Industry scoring failed: {e}")

        # ---- Spaltenreihenfolge: Score sichtbar + 52W-Spalten nach Close ----
    # score kommt aus dem Screener; wir nehmen ihn explizit nach vorne
    preferred_order = [
        "Company",
        "SA",
        "Industry",
        "Industry Ranking",
        "Industry Score",
        "Industry RS Score",
        "Industry Strong Stock Score",
        "Industry Volume Score",
        "MarketCap (Mio USD)",
        "EPS (Forward/TTM)",
        "EPS Wachstum FWD/TTM (%)",
        "EPS Wachstum letztes Q YoY (%)",
        "Revenue Wachstum TTM YoY (%)",
        "ROE (%)",
        "Operating Margin (%)",
        "FCF Margin (%)",
        "Debt to Equity",
        "EPS Acceleration (pp)",
        "Close Vorwoche",
        "Close", 
        "Veränderung in %",        
        "52W High",
        "Dist to 52W High (%)",
        "Ø-Volume 20T",
        "Volume Score",
        "RS (O'Neil)",
        "ΔRS 4W",
        "score",
        "SMA10W steigend",
        "SMA30W steigend",
        "SMA40W steigend",
        "MA-Ordnung 10>30>40",
        "52W Range OK",
        "RS-Trend ↑",
        "Vol-Breakout",
        "Close > Vorwoche",
        "MACD Bullish Cross (W)",
        "VCP",
        "VCP Waves",
        "VCP Entry",
        "VCP Breakout Level",
        "Launchpad",
        "Launchpad Score",
        "Launchpad Weeks",
        "Launchpad Range (%)",
        "Launchpad Pivot",
    ]

    existing_pref = [c for c in preferred_order if c in leaders.columns]
    remaining     = [c for c in leaders.columns if c not in existing_pref]
    leaders       = leaders[existing_pref + remaining]

    # ── Alpaca: nicht ausgelöste Orders der Vorwoche cancellen ───────────────
    if not TEST_MODE:
        cancelled = alpaca_client.cancel_open_orders()
        if cancelled > 0:
            print(f"[ORDER] 🗑  {cancelled} offene Order(s) aus der Vorwoche gecancelt")
    else:
        print("[ORDER] TEST-MODUS — cancel_open_orders übersprungen")

    # ── Alpaca: verfügbares Kapital & offene Positionen ──────────────────────
    alpaca_portfolio = alpaca_client.get_portfolio()
    if alpaca_portfolio is not None:
        alpaca_cash      = alpaca_portfolio["cash"]
        alpaca_positions = [p["symbol"] for p in alpaca_portfolio["positions"]]
        print(f"[ALPACA] Cash: ${alpaca_cash:,.0f} | Equity: ${alpaca_portfolio['equity']:,.0f} | Positionen: {alpaca_positions or '–'}")
    else:
        alpaca_cash      = None
        alpaca_positions = []
        print("[ALPACA] Nicht verbunden – Fallback auf account_equity aus Einstellungen")

    # ── Exit-Manager: MACD Bearish Cross auf offenen Positionen prüfen ──────────
    exit_results: list[dict] = []
    if alpaca_portfolio is not None:
        exit_results = exit_manager.run_exit_checks(alpaca_portfolio, dry_run=TEST_MODE)
        for r in exit_results:
            sym = r["symbol"]
            if r["cross"]:
                print(
                    f"[EXIT] 🔔 {sym}  MACD Bearish Cross — "
                    f"MACD={r['macd']}  Signal={r['signal']}  "
                    f"Raised Stop: ${r['new_stop']}  → {r['status']}"
                )
            else:
                print(
                    f"[EXIT] ✅ {sym}  kein Cross "
                    f"(MACD={r['macd']}, Signal={r['signal']})"
                )
    else:
        print("[EXIT] Kein Alpaca-Portfolio — Exit-Check übersprungen")

    # ── Trade Journal: Positionen synchronisieren (vor Signal-Generator) ────────
    journal_data: dict       = {}
    pt_results:   list[dict] = []
    if alpaca_portfolio is not None:
        filled_buys  = alpaca_client.get_filled_orders("buy")
        filled_sells = alpaca_client.get_filled_orders("sell")
        journal_data = trade_journal.sync(alpaca_portfolio, filled_buys, filled_sells)
        if exit_results:
            journal_data = trade_journal.apply_raised_stops(journal_data, exit_results)

        # ── Profit-Taking: Teilverkäufe und Stop-Nachzug ──────────────────────
        pt_results   = exit_manager.run_profit_taking_checks(journal_data, dry_run=TEST_MODE)
        journal_data = trade_journal.apply_profit_taking(journal_data, pt_results)
        if not any(r.get("actions_taken") for r in pt_results):
            print("[PROFIT] Keine Gewinnmitnahme-Aktionen diese Woche")

        trades_html = trade_journal.build_and_save_html(journal_data)
        open_count   = len(journal_data.get("open",   []))
        closed_count = len(journal_data.get("closed", []))
        print(f"[JOURNAL] {open_count} offen / {closed_count} geschlossen → {trades_html}")

        # ── Performance Dashboard ─────────────────────────────────────────────
        import portfolio_performance
        port_history = alpaca_client.get_portfolio_history()
        portfolio_performance.build_and_save(port_history)
    else:
        print("[JOURNAL] Kein Alpaca-Portfolio — Journal-Sync übersprungen")
        import portfolio_performance
        portfolio_performance.build_and_save()

    # ── Projected Cash: Erlöse aus Gewinnmitnahmen in Signal-Sizing einrechnen ─
    # Sell-Orders füllen sich erst Montag. Wir projizieren das erwartete Cash,
    # damit der Signal-Generator die korrekte Positionsgröße berechnen kann.
    # Die Kauforder wird erst nach Fill-Bestätigung platziert (--monday-execute).
    sell_symbols:  list[str] = []
    sell_proceeds: float     = 0.0
    for r in pt_results:
        actions      = r.get("actions_taken", [])
        real_actions = [a for a in actions if "dry" not in a]
        if not real_actions:
            continue
        sym   = r["symbol"]
        trade = next((t for t in journal_data.get("open", []) if t["symbol"] == sym), None)
        if trade:
            price = float(trade.get("current_price") or 0.0)
            if "partial_1" in real_actions:
                sell_proceeds += r.get("partial_sell_1_qty", 0) * price
                if sym not in sell_symbols:
                    sell_symbols.append(sym)
            if "partial_2" in real_actions:
                sell_proceeds += r.get("partial_sell_2_qty", 0) * price
                if sym not in sell_symbols:
                    sell_symbols.append(sym)

    projected_cash = (alpaca_cash or 0.0) + sell_proceeds
    if sell_proceeds > 0:
        print(
            f"[ORDER] 💰 Erwartete Sell-Erlöse: ${sell_proceeds:,.0f}  → "
            f"Projected Cash für Signal-Sizing: ${projected_cash:,.0f}"
        )

    # ── Trade-Signal-Generator (Blueprint-Regelwerk) ──────────────────────────
    signals, _signal_candidates, sector_excluded = generate_signals(
        leaders,
        market_bullish  = market_bullish,
        account_equity  = SETTINGS.account_equity,
        win_rate        = SETTINGS.win_rate,
        win_loss_ratio  = SETTINGS.win_loss_ratio,
        kelly_fraction  = SETTINGS.kelly_fraction,
        max_positions   = SETTINGS.max_positions,
        rules           = {"max_industry_rank": SETTINGS.max_industry_rank},
        available_cash  = projected_cash if projected_cash > 0 else alpaca_cash,
        open_positions  = alpaca_positions,
    )
    print(f"[SIGNALS] {len(signals)} Kaufsignal(e) gefunden")

    out_dir = Path("artifacts")
    out_dir.mkdir(parents=True, exist_ok=True)
    signals_json = save_signals_json(signals, out_dir / f"signals_{report_date}.json")
    print(f"[SIGNALS] Signale gespeichert → {signals_json}")

    # Persist signal metadata in docs/data/ so it survives across CI runs (committed to git)
    meta_json = save_signals_json(signals, Path("docs/data") / f"signals_meta_{report_date}.json")
    print(f"[SIGNALS] Signal-Metadaten persistiert → {meta_json}")

    # ── Alpaca: OTO Orders sofort platzieren ODER als Pending zurückhalten ─────
    top_picks = [s for s in signals if s.is_top_pick]
    if sell_symbols and top_picks and alpaca_portfolio is not None:
        # Sell-Orders wurden gerade platziert → Käufe erst nach Fill-Bestätigung
        _save_pending_orders(top_picks, sell_symbols, sell_proceeds, report_date)
        print(
            f"[ORDER] ⏳ {len(top_picks)} Kauforder(s) zurückgehalten — "
            f"warte auf Sell-Bestätigung ({', '.join(sell_symbols)})"
        )
        print("[ORDER]    Montag ausführen: python main.py --monday-execute")
    elif signals and alpaca_portfolio is not None:
        order_results = alpaca_client.place_signal_orders(signals, dry_run=TEST_MODE)
        for r in order_results:
            status = r["status"]
            ticker = r["ticker"]
            if status == "placed":
                print(
                    f"[ORDER] ✅ {ticker}  Buy-Stop ${r['buy_stop']}  "
                    f"Max-Gap ${r['max_gap']}  Stop ${r['stop_loss']}  "
                    f"Qty {r['qty']}  → Order-ID {r['order_id']}"
                )
            elif status.startswith("skip"):
                print(f"[ORDER] ⏭  {ticker}: {status}")
            elif status == "dry_run":
                print(f"[ORDER] 🔍 DRY-RUN {ticker}  qty={r['qty']}")
            else:
                print(f"[ORDER] ❌ {ticker}: {status}")
    else:
        print("[ORDER] Keine Orders platziert (kein Alpaca-Client oder keine Signale)")

    # --- Formatierte Kopie NUR für HTML-Report ---
    leaders_html = leaders.copy()

    def fmt_2dec(x):
        return f"{x:.2f}" if pd.notna(x) else "–"

    def fmt_int(x):
        return f"{x:,.0f}" if pd.notna(x) else "–"

    def fmt_0dec(x):
        return f"{x:.0f}" if pd.notna(x) else "–"
    
    # Spalten mit 2 Nachkommastellen
    for col in [
        "Industry Score",
        "Industry RS Score",
        "Industry Strong Stock Score",
        "Industry Volume Score",
        "EPS (Forward/TTM)",
        "EPS Wachstum FWD/TTM (%)",
        "EPS Wachstum letztes Q YoY (%)",
        "Revenue Wachstum TTM YoY (%)",
        "ROE (%)",
        "Operating Margin (%)",
        "FCF Margin (%)",
        "Debt to Equity",
        "EPS Acceleration (pp)",
        "Close",
        "Close Vorwoche",
        "Veränderung in %",
        "52W High",
        "Dist to 52W High (%)",
        "Volume Score",
        "ΔRS 4W",
        "VCP Entry",
        "VCP Breakout Level",
        "ATR / Price (%)",
        "Launchpad Range (%)",
        "Launchpad Pivot",
    ]:
        if col in leaders_html.columns:
            leaders_html[col] = leaders_html[col].apply(fmt_2dec)

    # Spalten mit ganzen Zahlen
    for col in [
        "MarketCap (Mio USD)",
        "Ø-Volume 20T",
    ]:
        if col in leaders_html.columns:
            leaders_html[col] = leaders_html[col].apply(fmt_int)

    # RS als ganze Zahl formatieren
    if "RS (O'Neil)" in leaders_html.columns:
        leaders_html["RS (O'Neil)"] = leaders_html["RS (O'Neil)"].apply(fmt_0dec)

    #Screener-Ausgabe prüfen
    print(f"[DEBUG] Found {len(leaders)} Minervini leaders")

    # ── GitHub Pages: vollständigen Report speichern ──────────────────────────
    PAGES_BASE_URL = "https://weekly-market-condition.pages.dev"
    report_url     = f"{PAGES_BASE_URL}/reports/{report_date}.html"

    html_full = build_html_report(
        breadth_df, idx_df, risk_df, summary, report_date,
        weekly, leaders_html, signals=signals, pages_url=None,
        alpaca_cash=alpaca_cash, alpaca_positions=alpaca_positions, alpaca_portfolio=alpaca_portfolio,
        sector_excluded=sector_excluded,
        sp500_breadth_pct=sp500_breadth_pct, min_breadth_pct=_min_breadth,
        test_mode=TEST_MODE, sector_rows=sector_rows,
    )
    docs_reports_dir = Path("docs/reports")
    docs_reports_dir.mkdir(parents=True, exist_ok=True)
    report_file = docs_reports_dir / f"{report_date}.html"
    report_file.write_text(html_full, encoding="utf-8")
    print(f"[PAGES] Report gespeichert → {report_file}")

    # Index-Seite aktualisieren
    from report_builder import build_index_page
    index_path = Path("docs/index.html")
    index_path.write_text(
        build_index_page(docs_reports_dir, PAGES_BASE_URL),
        encoding="utf-8",
    )
    print(f"[PAGES] Index aktualisiert → {index_path}")

    # ── E-Mail: kompakte Version ohne große Signaltabelle ─────────────────────
    html_email = build_html_report(
        breadth_df, idx_df, risk_df, summary, report_date,
        weekly, leaders_html, signals=signals, pages_url=report_url,
        alpaca_cash=alpaca_cash, alpaca_positions=alpaca_positions, alpaca_portfolio=alpaca_portfolio,
        sector_excluded=sector_excluded,
        sp500_breadth_pct=sp500_breadth_pct, min_breadth_pct=_min_breadth,
        test_mode=TEST_MODE, sector_rows=sector_rows,
    )

    # E-Mail Betreff zeigt Signalanzahl + TEST-MODUS-Hinweis
    signal_count = len(signals)
    tm_prefix = "[TEST-MODUS] " if TEST_MODE else ""
    email_subject = f"{tm_prefix}Weekly US Market Report — {signal_count} Kaufsignal{'e' if signal_count != 1 else ''}"

    # ── Früher Rücksprung wenn Excel-Export deaktiviert (Standard) ────────────
    if not SETTINGS.export_excel:
        send_email(html_email, subject_suffix=email_subject, attachments=None)
        return

    # ── Ab hier: Excel-Export (nur wenn EXPORT_EXCEL=true) ───────────────────
    leaders_out = leaders.reset_index().rename(columns={"index": "Ticker"})
    out_path = out_dir / f"market_leaders_{report_date}.xlsx"
    
    # 2) Immer schreiben – auch wenn leer (dann gibt's wenigstens Header)
    # Excel Export:
    # - Sheet 'Leaders' enthält nur Industry Ranking + Industry RS Score als Industry-Metriken
    # - Sheet 'Industries' enthält alle Industry-Metriken (Ranking + Teilmetriken + Composite)
    leaders_out_excel = leaders_out.copy()
    drop_ind_cols = [
        'Industry Score',
        'Industry Strong Stock Score',
        'Industry Volume Score',
    ]
    leaders_out_excel.drop(columns=[c for c in drop_ind_cols if c in leaders_out_excel.columns], inplace=True, errors='ignore')

    # ------------------------------------------------------------
    # Industry-relative percentiles for coloring ROE / Margins
    # (used only for Excel conditional formatting)
    # ------------------------------------------------------------
    if "Industry" in leaders_out_excel.columns:
        for metric, pctl_col in [
            ("ROE (%)", "ROE Ind Pctl"),
            ("Operating Margin (%)", "OpMargin Ind Pctl"),
            ("FCF Margin (%)", "FCFMargin Ind Pctl"),
        ]:
            if metric in leaders_out_excel.columns:
                leaders_out_excel[pctl_col] = (
                    leaders_out_excel.groupby("Industry")[metric]
                    .transform(lambda s: pd.to_numeric(s, errors="coerce").rank(pct=True))
                )
    
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        leaders_out_excel.to_excel(writer, index=False, sheet_name='Leaders')
        # Industries sheet (may be empty if scoring failed)
        if industry_table is not None and not industry_table.empty:
            # Sort industries by ranking (ascending: 1 is best)
            industry_out = industry_table.copy()
            if 'Industry Ranking' in industry_out.columns:
                industry_out = industry_out.sort_values('Industry Ranking', ascending=True, kind='mergesort')
            # Ensure column order (Industry, Sektor, ...)
            if 'Sektor' in industry_out.columns:
                cols = ['Industry', 'Sektor'] + [c for c in industry_out.columns if c not in ('Industry','Sektor')]
                industry_out = industry_out[cols]
            industry_out.to_excel(writer, index=False, sheet_name='Industries')
        else:
            # Write an empty template so the sheet always exists
            pd.DataFrame(columns=['Industry', 'Sektor', 'Industry Ranking', 'Industry_RS_raw', 'Valid_RS_Count',
                                  'Industry RS Score', 'Industry Strong Stock Score', 'Activity', 'Direction',
                                  'Industry Volume Score', 'Industry Score']).to_excel(
                writer, index=False, sheet_name='Industries'
            )

    
    # Excel laden - neu
    wb = load_workbook(out_path)
    ws = wb['Leaders']

    # Spalte "SA" finden
    sa_col_idx = None
    for cell in ws[1]:
        if cell.value == "SA":
            sa_col_idx = cell.column
            break
    
    if sa_col_idx is not None:
        sa_col_letter = get_column_letter(sa_col_idx)
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{sa_col_letter}{row}"]
            url = cell.value
            if url and isinstance(url, str):
                cell.value = "SA"
                cell.hyperlink = url
                cell.font = Font(color="0000EE", underline="single")  # Blau + Unterstrichen
    
    # -------------------------------
    # Auto-Fit für alle Spalten
    # -------------------------------
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
    
    # -------------------------------
    # Number-Format Regeln
    # -------------------------------
    
    # Spaltennamen zu Spaltenindex mappen
    header_row = {cell.value: cell.column for cell in ws[1] if cell.value}
    
    # 2 Nachkommastellen
    two_dec_cols = [
        "Industry Score",
        "Industry RS Score",
        "Industry Strong Stock Score",
        "Industry Volume Score",
        "EPS (Forward/TTM)",
        "EPS Wachstum FWD/TTM (%)",
        "EPS Wachstum letztes Q YoY (%)",
        "Revenue Wachstum TTM YoY (%)",
        "ROE (%)",
        "Operating Margin (%)",
        "FCF Margin (%)",
        "Debt to Equity",
        "EPS Acceleration (pp)",
        "Close",
        "Close Vorwoche",
        "Veränderung in %",
        "52W High",
        "Dist to 52W High (%)",
        "VCP Breakout Level",
        "Volume Score",
        "ATR / Price (%)",
        "Launchpad Range (%)",
        "Launchpad Pivot",
    ]
    
    # Ganze Zahlen (ohne Nachkommastellen)
    zero_dec_cols = [
        "MarketCap (Mio USD)",
        "Ø-Volume 20T",
    ]
    
    # --- Anwenden der Formate ---
    for col_name in two_dec_cols:
        if col_name in header_row:
            col_letter = get_column_letter(header_row[col_name])
            for cell in ws[col_letter][1:]:  # alle Zeilen außer Header
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"
    
    for col_name in zero_dec_cols:
        if col_name in header_row:
            col_letter = get_column_letter(header_row[col_name])
            for cell in ws[col_letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
                    
    # 
    # --- Hintergrundfarben für Fundamentals wie im ursprünglichen Report ---
    # Debt-to-Equity & EPS Acceleration: banded thresholds
    apply_debt_eps_conditional_formatting(ws, debt_header="Debt to Equity", eps_header="EPS Acceleration (pp)", start_row=2)

    # ROE / Margins: industry-relative percentile coloring (helper columns are hidden)
    apply_industry_percentile_conditional_formatting(ws, metric_header="ROE (%)", pctl_header="ROE Ind Pctl", start_row=2, hide_pctl=True)
    apply_industry_percentile_conditional_formatting(ws, metric_header="Operating Margin (%)", pctl_header="OpMargin Ind Pctl", start_row=2, hide_pctl=True)
    apply_industry_percentile_conditional_formatting(ws, metric_header="FCF Margin (%)", pctl_header="FCFMargin Ind Pctl", start_row=2, hide_pctl=True)

    # --- Boolesche Spalten (Minervini-Kriterien) einfärben ---
    style_boolean_columns(ws)

    # -------------------------------------------------------------------------
    # LAUNCHPAD SCORE - Gradient Conditional Formatting (90+ = Excellent)
    # -------------------------------------------------------------------------
    # Verwendet _apply_cf_formula_fill analog zu Debt/EPS
    
    lp_score_col = _col_letter_by_header(ws, "Launchpad Score")
    
    if lp_score_col:
        rng = f"{lp_score_col}2:{lp_score_col}{ws.max_row}"
        c = lp_score_col
        r = 2
        
        # NaN/Empty: Grau
        _apply_cf_formula_fill(
            ws, rng, 
            f"NOT(ISNUMBER(${c}{r}))", 
            CF_GRAY_RGB, 
            stop=False
        )
        
        # Score >= 90: Dunkelgrün (Excellent)
        _apply_cf_formula_fill(
            ws, rng, 
            f"AND(ISNUMBER(${c}{r}),${c}{r}>=90)", 
            "00B050",  # Dark Green
            stop=True
        )
        
        # Score 80-89: Hellgrün (Very Good)
        _apply_cf_formula_fill(
            ws, rng, 
            f"AND(ISNUMBER(${c}{r}),${c}{r}>=80,${c}{r}<90)", 
            CF_GREEN_RGB,
            stop=True
        )
        
        # Score 70-79: Gelb (OK)
        _apply_cf_formula_fill(
            ws, rng, 
            f"AND(ISNUMBER(${c}{r}),${c}{r}>=70,${c}{r}<80)", 
            CF_YELLOW_RGB,
            stop=True
        )
        
        # Score < 70: Grau (Weak)
        _apply_cf_formula_fill(
            ws, rng, 
            f"AND(ISNUMBER(${c}{r}),${c}{r}<70)", 
            CF_GRAY_RGB,
            stop=True
        )

    # -------------------------------------------------------------------------
    # LAUNCHPAD RANGE % - Tighter = Better (< 8% = Green)
    # -------------------------------------------------------------------------
    
    lp_range_col = _col_letter_by_header(ws, "Launchpad Range (%)")
    
    if lp_range_col:
        rng = f"{lp_range_col}2:{lp_range_col}{ws.max_row}"
        c = lp_range_col
        r = 2
        
        # NaN/Empty: Grau
        _apply_cf_formula_fill(
            ws, rng, 
            f"NOT(ISNUMBER(${c}{r}))", 
            CF_GRAY_RGB, 
            stop=False
        )
        
        # Range < 8%: Grün (Sehr tight)
        _apply_cf_formula_fill(
            ws, rng, 
            f"AND(ISNUMBER(${c}{r}),${c}{r}<8)", 
            CF_GREEN_RGB,
            stop=True
        )
        
        # Range 8-10%: Gelb (OK)
        _apply_cf_formula_fill(
            ws, rng, 
            f"AND(ISNUMBER(${c}{r}),${c}{r}>=8,${c}{r}<10)", 
            CF_YELLOW_RGB,
            stop=True
        )
        
        # Range >= 10%: Orange (Zu breit)
        _apply_cf_formula_fill(
            ws, rng, 
            f"AND(ISNUMBER(${c}{r}),${c}{r}>=10)", 
            CF_ORANGE_RGB,
            stop=True
        )

    # -------------------------------------------------------------------------
    # BONUS: GOLD BORDER für Stocks mit BEIDEN Patterns (VCP + Launchpad)
    # -------------------------------------------------------------------------
    # Highlightet Zeilen wo sowohl VCP=True als auch Launchpad=True
    
    vcp_col_idx = header_row.get("VCP")
    lp_col_idx = header_row.get("Launchpad")
    
    if vcp_col_idx and lp_col_idx:
        from openpyxl.styles import Border, Side
        
        gold_border = Border(
            top=Side(style='thick', color='FF6600'),
            bottom=Side(style='thick', color='FF6600')
        )
        
        for row in range(2, ws.max_row + 1):
            vcp_val = ws.cell(row=row, column=vcp_col_idx).value
            lp_val = ws.cell(row=row, column=lp_col_idx).value
            
            # Beide Patterns erkannt
            if vcp_val in (True, "True", "TRUE", 1) and lp_val in (True, "True", "TRUE", 1):
                # Dickeren Border für die gesamte Zeile
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = gold_border

    # -------------------------------------------------------------------------
    # END OF LAUNCHPAD FORMATTING
    # -------------------------------------------------------------------------

    

    # -------------------------------
    # Format Industries sheet
    # -------------------------------
    if 'Industries' in wb.sheetnames:
        ws_ind = wb['Industries']

        # Auto-width based on header length (not data length)
        for col_idx, cell in enumerate(ws_ind[1], start=1):
            header = "" if cell.value is None else str(cell.value)
            ws_ind.column_dimensions[get_column_letter(col_idx)].width = max(10, len(header) + 2)

        # Number formats (2 decimals) for selected industry metrics
        header_ind = {cell.value: cell.column for cell in ws_ind[1] if cell.value}
        ind_two_dec_cols = [
            "Industry RS Score",
            "Industry Strong Stock Score",
            "Activity",
            "Direction",
            "Industry Volume Score",
            "Industry Score",
        ]
        for col_name in ind_two_dec_cols:
            if col_name in header_ind:
                col_letter = get_column_letter(header_ind[col_name])
                for cell in ws_ind[col_letter][1:]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.00"

        # Keep ranking as integer
        if "Industry Ranking" in header_ind:
            col_letter = get_column_letter(header_ind["Industry Ranking"])
            for cell in ws_ind[col_letter][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0"

    wb.save(out_path)
    
    # 4) Beim Mailversand denselben Pfad anhängen
    send_email(html_email, subject_suffix=email_subject, attachments=[str(out_path)])

if __name__ == "__main__":
    if "--monday-execute" in sys.argv:
        monday_execute()
    else:
        run()
