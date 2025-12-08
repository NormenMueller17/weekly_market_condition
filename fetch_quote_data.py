import os
import time
import pandas as pd
from datetime import datetime
from config import SETTINGS
from data_sources import get_universe, get_company_info_map_from_csv, load_weekly_history, load_index_series
from breadth import compute_breadth, compute_breadth_snapshots_with_advancers as compute_breadth_snapshots
from emailer import send_email
from screener import screen_universe_minervini
from fetch_quote_data import batch_fetch_quote_data, fetch_quote_data_single
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import yfinance as yf

def batch_fetch_quote_data(tickers) -> dict:
	"""
	Holt Fundamentaldaten für eine Liste von Tickern parallel.
	Rückgabe: dict[ticker] -> dict mit Feldern wie in fetch_quote_data_single.
	"""
	results = {}
	tickers = list(dict.fromkeys(tickers))  # Duplikate raus

	# Anzahl Threads begrenzen – 8-16 ist ein guter Startwert
	max_workers = min(16, max(4, len(tickers) // 20))  # z.B. 4–16 Threads

	print(f"[INFO] Starte batch_fetch_quote_data für {len(tickers)} Ticker "
		  f"mit {max_workers} Threads ...")

	with ThreadPoolExecutor(max_workers=max_workers) as executor:
		future_to_ticker = {executor.submit(fetch_quote_data_single, t): t for t in tickers}

		for future in as_completed(future_to_ticker):
			tkr = future_to_ticker[future]
			try:
				data = future.result()
			except Exception as e:
				print(f"[ERROR] batch_fetch_quote_data: {tkr} -> {e}")
				data = {
					"Close": None,
					"MarketCap_Mio": None,
					"EPS_FWD_TTM": None,
					"EPS_GROWTH_FWD_TTM": None,
					"REV_GROWTH_TTM_YOY": None,
				}
			results[tkr] = data
	return results

    # --- Aktuellen Schlusskurs & Marktkapitalisierung ergänzen ---
    def fetch_quote_data_single(ticker: str) -> dict:
        """
        Holt Close, MarketCap (Mio), EPS (Forward/TTM) und Revenue Growth (TTM YoY)
        für EINEN Ticker. Mit kleinem Retry-Mechanismus.
        """
        MAX_RETRIES = 3
        SLEEP_SECONDS = 0.75
    
        for attempt in range(MAX_RETRIES):
            try:
                info = yf.Ticker(ticker)
                fast = getattr(info, "fast_info", {}) or {}
    
                # Close
                close = (
                    fast.get("lastPrice")
                    or fast.get("last_price")
                    or info.info.get("regularMarketPrice")
                )
    
                # MarketCap
                market_cap = fast.get("marketCap") or info.info.get("marketCap")
                market_cap_mio = market_cap / 1_000_000 if market_cap else None
    
                # EPS: Forward + TTM
                eps_forward = (
                    fast.get("epsForward")
                    or info.info.get("forwardEps")
                )
                eps_trailing = (
                    fast.get("epsTrailingTwelveMonths")
                    or info.info.get("trailingEps")
                )
                eps_fwd_ttm = eps_forward if eps_forward is not None else eps_trailing
    
                # EPS-Wachstum (Forward vs. TTM)
                eps_growth_pct = None
                if eps_forward is not None and eps_trailing not in (None, 0):
                    try:
                        eps_growth_pct = (eps_forward / eps_trailing - 1.0) * 100.0
                    except Exception:
                        eps_growth_pct = None
    
                # Revenue Growth (TTM YoY)
                rev_growth_pct = None
                try:
                    rg = info.info.get("revenueGrowth")
                    if rg is not None:
                        rev_growth_pct = float(rg) * 100.0
                except Exception:
                    rev_growth_pct = None
    
                return {
                    "Close": close,
                    "MarketCap_Mio": market_cap_mio,
                    "EPS_FWD_TTM": eps_fwd_ttm,
                    "EPS_GROWTH_FWD_TTM": eps_growth_pct,
                    "REV_GROWTH_TTM_YOY": rev_growth_pct,
                }
    
            except Exception as e:
                if attempt < MAX_RETRIES - 1:
                    print(f"[WARN] fetch_quote_data_single({ticker}) Versuch {attempt+1} fehlgeschlagen ({e}). "
                          f"Retry in {SLEEP_SECONDS}s ...")
                    time.sleep(SLEEP_SECONDS)
                    continue
                else:
                    print(f"[ERROR] fetch_quote_data_single({ticker}) dauerhaft fehlgeschlagen ({e}).")
    
        # Fallback: alles None
        return {
            "Close": None,
            "MarketCap_Mio": None,
            "EPS_FWD_TTM": None,
            "EPS_GROWTH_FWD_TTM": None,
            "REV_GROWTH_TTM_YOY": None,
        }
